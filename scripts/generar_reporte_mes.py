"""
================================================================
  RUMBO VERDE — GENERADOR DE REPORTE MENSUAL AUTOMÁTICO
  Versión: 2.0 | Integrado con Matriz Maestra
================================================================

USO:
    python generar_reporte_mes.py mensuales/2026-06_ventas.xlsx

QUÉ HACE:
    1. Lee el archivo del mes nuevo
    2. Actualiza la matriz maestra con el nuevo mes
    3. Genera reporte PDF con:
       - KPIs del mes actual
       - Comparativo vs mismo mes año anterior (YoY)
       - Comparativo vs mes anterior (MoM)
       - Tendencia diaria + contexto histórico
       - Mix de canales
       - Top productos y marcas con Pareto
       - Insights automáticos
    4. Exporta resumen Excel del mes

ESTRUCTURA DE CARPETAS EN TU MAC:
    ventas_rumbo_verde/
    ├── anuales/           → archivos anuales históricos
    ├── mensuales/         → un Excel por mes
    ├── reportes/          → PDFs generados (se crea automático)
    ├── consolidar_matriz.py
    └── generar_reporte_mes.py  ← este script
================================================================
"""

import sys
import os
import re
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyBboxPatch
import matplotlib.ticker as mticker
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ================================================================
#  CONFIGURACIÓN
# ================================================================
NOMBRE_TIENDA    = "Rumbo Verde"
CARPETA_ANUALES  = os.path.expanduser("~/Documents/Rumbo Verde/Ventas/anuales")
CARPETA_MENSUALES= os.path.expanduser("~/Documents/Rumbo Verde/Ventas/mensuales")
CARPETA_REPORTES = os.path.expanduser("~/Documents/Rumbo Verde/Reportes")
MATRIZ_PATH      = os.path.expanduser("~/Documents/Ciencia de Datos/rumbo-verde-analisis/matriz/matriz_maestra.xlsx")

COLORES = {
    "primario":    "#1B4332",
    "secundario":  "#40916C",
    "acento":      "#52B788",
    "claro":       "#D8F3DC",
    "alerta":      "#E76F51",
    "dorado":      "#F4A261",
    "azul":        "#1A759F",
    "neutro":      "#F8FAF8",
    "texto":       "#1B1B2F",
    "fondo":       "#FFFFFF",
    "borde":       "#D0E8D8",
    "positivo":    "#2D6A4F",
    "negativo":    "#AE2012",
}

CANAL_MAP = {
    "lista de precios base rv final": "Tienda / Web",
    "lista de precios base rv ":      "Tienda / Web",
    "cafeteria rumbo verde":          "Tienda / Web",
    "precio mayorista":               "Mayorista",
    "cornershop":                     "Cornershop",
    "uber":                           "Uber Eats",
    "rappi":                          "Rappi",
    "ml full":                        "Mercado Libre",
}

MESES_ES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
            7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}


# ================================================================
#  PASO 1 — CARGAR Y LIMPIAR EL MES NUEVO
# ================================================================
def normalizar_canal(lista_precio):
    if pd.isna(lista_precio): return "Tienda / Web"
    lp = str(lista_precio).lower().strip()
    for key, canal in CANAL_MAP.items():
        if key in lp: return canal
    return "Tienda / Web"


def cargar_mes(path: str) -> tuple:
    """Carga el archivo mensual y retorna df_prod, df_desp, año, mes."""
    print(f"\n📂 Cargando: {os.path.basename(path)}")
    df = pd.read_excel(path)
    df["Fecha Venta"] = pd.to_datetime(df["Fecha Venta"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["Fecha Venta"])
    df["año"] = df["Fecha Venta"].dt.year
    df["mes"]  = df["Fecha Venta"].dt.month
    df["dia"]  = df["Fecha Venta"].dt.day
    df["dia_semana"] = df["Fecha Venta"].dt.day_name()
    df["canal"] = df["Lista de Precio"].apply(normalizar_canal)

    año = int(df["año"].mode()[0])
    mes = int(df["mes"].mode()[0])

    df = df[(df["año"] == año) & (df["mes"] == mes)]
    df_prod = df[(df["Tipo Movimiento"] == "venta") & (df["Costo neto unitario"] > 0)].copy()
    df_desp = df[(df["Tipo Movimiento"] == "venta") & (df["Costo neto unitario"] == 0)].copy()
    df_devs = df[df["Tipo Movimiento"] == "devolucion"].copy()

    # Excluir anomalías de costo
    anomalo = (df_prod["Costo neto unitario"] > df_prod["Precio Bruto Unitario"] * 2) & \
              (df_prod["Precio Bruto Unitario"] < 100)
    df_prod = df_prod[~anomalo].copy()

    print(f"   ✅ {año}-{mes:02d} | {len(df_prod):,} líneas | "
          f"{df_prod['Tracking number'].nunique():,} boletas | "
          f"${df_prod['Venta Total Bruta'].sum()/1e6:.2f}M")
    return df_prod, df_desp, df_devs, año, mes


# ================================================================
#  PASO 2 — CALCULAR KPIs DEL MES
# ================================================================
def calcular_kpis_mes(df_prod, df_desp, df_devs) -> dict:
    """Calcula todos los KPIs del mes."""
    venta_bruta  = df_prod["Venta Total Bruta"].sum()
    venta_neta   = df_prod["Venta Total Neta"].sum()
    margen_bruto = df_prod["Margen"].sum()
    pct_margen   = margen_bruto / venta_neta if venta_neta > 0 else 0
    tickets      = df_prod["Tracking number"].nunique()
    unidades     = df_prod["Cantidad"].sum()
    boleta_prom  = venta_bruta / tickets if tickets > 0 else 0
    u_ticket     = unidades / tickets if tickets > 0 else 0
    precio_art   = venta_bruta / unidades if unidades > 0 else 0
    dias_op      = df_prod["dia"].nunique()
    venta_diaria = venta_bruta / dias_op if dias_op > 0 else 0

    # Descuentos
    monto_dto      = df_prod["Descuento Bruto"].sum()
    pct_lineas_dto = (df_prod["% Descuento"] > 0).sum() / len(df_prod) * 100

    # Despachos y devoluciones
    ingr_despacho = df_desp["Venta Total Bruta"].sum()
    monto_devs    = abs(df_devs["Venta Total Bruta"].sum()) if len(df_devs) > 0 else 0

    # Tendencia diaria
    por_dia = df_prod.groupby("dia").agg(
        ventas=("Venta Total Bruta","sum"),
        tickets=("Tracking number","nunique")
    ).reset_index()

    # Día de semana
    orden_dias = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    nombres_es = {"Monday":"Lunes","Tuesday":"Martes","Wednesday":"Miércoles",
                  "Thursday":"Jueves","Friday":"Viernes","Saturday":"Sábado","Sunday":"Domingo"}
    por_dia_sem = df_prod.groupby("dia_semana")["Venta Total Bruta"].sum().reindex(orden_dias).fillna(0)
    por_dia_sem.index = [nombres_es[d] for d in por_dia_sem.index]

    # Canales
    por_canal = df_prod.groupby("canal")["Venta Total Bruta"].sum().sort_values(ascending=False)
    venta_tienda = por_canal.get("Tienda / Web", 0)
    venta_uber   = por_canal.get("Uber Eats", 0)
    pct_tienda   = venta_tienda / venta_bruta * 100 if venta_bruta > 0 else 0
    pct_uber     = venta_uber   / venta_bruta * 100 if venta_bruta > 0 else 0

    # Pareto marcas
    marcas_v = df_prod.groupby("Marca")["Venta Total Bruta"].sum().sort_values(ascending=False)
    marcas_v = marcas_v[marcas_v.index.notna()]
    acum_m = marcas_v.cumsum() / marcas_v.sum()
    marcas_80 = int((acum_m <= 0.80).sum()) + 1
    marca_1 = marcas_v.index[0] if len(marcas_v) > 0 else ""
    marca_1_pct = marcas_v.iloc[0] / venta_bruta * 100 if venta_bruta > 0 else 0

    # Top 10 productos
    top_prods = df_prod.groupby("Producto / Servicio")["Venta Total Bruta"].sum().sort_values(ascending=False).head(10)

    # Día estrella
    dia_max = por_dia.loc[por_dia["ventas"].idxmax()]

    return {
        "venta_bruta": venta_bruta,   "venta_neta": venta_neta,
        "margen_bruto": margen_bruto, "pct_margen": pct_margen,
        "tickets": tickets,           "unidades": unidades,
        "boleta_prom": boleta_prom,   "u_ticket": u_ticket,
        "precio_art": precio_art,     "dias_op": dias_op,
        "venta_diaria": venta_diaria,
        "monto_dto": monto_dto,       "pct_lineas_dto": pct_lineas_dto,
        "ingr_despacho": ingr_despacho, "monto_devs": monto_devs,
        "por_dia": por_dia,           "por_dia_sem": por_dia_sem,
        "por_canal": por_canal,       "pct_tienda": pct_tienda,
        "pct_uber": pct_uber,
        "marcas_80": marcas_80,       "marca_1": marca_1,
        "marca_1_pct": marca_1_pct,   "marcas_v": marcas_v,
        "top_prods": top_prods,
        "dia_max_num": int(dia_max["dia"]),
        "dia_max_venta": dia_max["ventas"],
    }


# ================================================================
#  PASO 3 — LEER CONTEXTO HISTÓRICO DESDE LA MATRIZ
# ================================================================
def leer_contexto_historico(año: int, mes: int) -> dict:
    """
    Lee la matriz maestra para obtener:
    - KPIs del mismo mes año anterior (YoY)
    - KPIs del mes anterior (MoM)
    - Promedio histórico del mes (todos los años)
    - Tendencia de los últimos 12 meses
    """
    if not os.path.exists(MATRIZ_PATH):
        print("  ⚠️  Matriz maestra no encontrada — sin contexto histórico")
        return {}

    wb = openpyxl.load_workbook(MATRIZ_PATH, data_only=True)
    ws = wb.active
    headers = [ws.cell(3, j).value for j in range(1, ws.max_column + 1)]
    rows = []
    for i in range(4, ws.max_row + 1):
        row = [ws.cell(i, j).value for j in range(1, ws.max_column + 1)]
        if any(v is not None for v in row):
            rows.append(row)

    df = pd.DataFrame(rows, columns=headers)

    def n(col):
        return pd.to_numeric(df[col], errors="coerce") if col in df.columns else pd.Series([np.nan]*len(df))

    df["_año"] = n("Año").astype("Int64")
    df["_mes"] = n("Mes").astype("Int64")
    df["_vb"]  = n("Venta Bruta ($)")
    df["_mg"]  = n("% Margen")
    df["_bol"] = n("Boletas")
    df["_bp"]  = n("Boleta Promedio ($)")
    df["_ut"]  = n("Unidades/Ticket")

    ctx = {}

    # Mismo mes año anterior
    ly = df[(df["_año"] == año - 1) & (df["_mes"] == mes)]
    if len(ly) > 0:
        ctx["ly_venta"]  = float(ly["_vb"].iloc[0]) if pd.notna(ly["_vb"].iloc[0]) else None
        ctx["ly_margen"] = float(ly["_mg"].iloc[0]) if pd.notna(ly["_mg"].iloc[0]) else None
        ctx["ly_boletas"]= int(ly["_bol"].iloc[0]) if pd.notna(ly["_bol"].iloc[0]) else None
        ctx["ly_bp"]     = float(ly["_bp"].iloc[0]) if pd.notna(ly["_bp"].iloc[0]) else None

    # Mes anterior
    mes_ant = mes - 1 if mes > 1 else 12
    año_ant = año if mes > 1 else año - 1
    pm = df[(df["_año"] == año_ant) & (df["_mes"] == mes_ant)]
    if len(pm) > 0:
        ctx["pm_venta"]  = float(pm["_vb"].iloc[0]) if pd.notna(pm["_vb"].iloc[0]) else None
        ctx["pm_boletas"]= int(pm["_bol"].iloc[0]) if pd.notna(pm["_bol"].iloc[0]) else None
        ctx["pm_bp"]     = float(pm["_bp"].iloc[0]) if pd.notna(pm["_bp"].iloc[0]) else None

    # Promedio histórico del mes (todos los años disponibles excepto el actual)
    hist = df[(df["_mes"] == mes) & (df["_año"] < año)]
    if len(hist) > 0:
        ctx["hist_venta_prom"]  = float(hist["_vb"].mean())
        ctx["hist_margen_prom"] = float(hist["_mg"].mean())
        ctx["hist_boletas_prom"]= float(hist["_bol"].mean())

    # Últimos 12 meses para mini tendencia
    df_sorted = df.sort_values(["_año","_mes"])
    ult12 = df_sorted[
        (df_sorted["_año"] * 100 + df_sorted["_mes"]) < (año * 100 + mes)
    ].tail(12)
    if len(ult12) > 0:
        ctx["ult12_periodos"] = [f"{int(r['_año'])}-{int(r['_mes']):02d}" for _, r in ult12.iterrows()]
        ctx["ult12_ventas"]   = [float(v) if pd.notna(v) else 0 for v in ult12["_vb"]]
        ctx["ult12_margenes"] = [float(v) if pd.notna(v) else 0 for v in ult12["_mg"]]

    return ctx


# ================================================================
#  PASO 4 — GENERAR INSIGHTS AUTOMÁTICOS
# ================================================================
def generar_insights(kpis: dict, ctx: dict, año: int, mes: int) -> list:
    insights = []

    # YoY ventas
    if ctx.get("ly_venta"):
        yoy = (kpis["venta_bruta"] - ctx["ly_venta"]) / ctx["ly_venta"] * 100
        emoji = "📈" if yoy >= 0 else "📉"
        dir_txt = "por encima" if yoy >= 0 else "por debajo"
        insights.append(
            f"{emoji} Ventas {dir_txt} del mismo mes {año-1} en {abs(yoy):.1f}% "
            f"(${kpis['venta_bruta']/1e6:.1f}M vs ${ctx['ly_venta']/1e6:.1f}M). "
            + ("Buen desempeño — mantener foco en mix de producto." if yoy >= 0
               else "Señal de alerta — revisar mix de canales y disponibilidad de stock.")
        )

    # YoY margen
    if ctx.get("ly_margen"):
        d_mg = kpis["pct_margen"] * 100 - ctx["ly_margen"]
        insights.append(
            f"{'✅' if d_mg >= 0 else '⚠️ '} Margen {kpis['pct_margen']*100:.1f}% "
            f"({'+'if d_mg>=0 else ''}{d_mg:.1f}pp vs {año-1}). "
            + ("El mix de productos está mejorando la rentabilidad." if d_mg >= 0
               else "Revisar si hay cambio en mix de canal o incremento de descuentos.")
        )

    # Boleta promedio
    if ctx.get("ly_bp"):
        d_bp = (kpis["boleta_prom"] - ctx["ly_bp"]) / ctx["ly_bp"] * 100
        insights.append(
            f"🛒 Boleta promedio ${kpis['boleta_prom']/1000:.0f}K "
            f"({'+'if d_bp>=0 else ''}{d_bp:.1f}% vs {año-1}). "
            + ("Clientes comprando más por visita — estrategia de cross-sell funcionando." if d_bp >= 0
               else "Posible presión de precios o cambio en perfil de cliente — monitorear.")
        )

    # Canales
    if kpis["pct_uber"] > 15:
        insights.append(
            f"🚀 Uber Eats representa el {kpis['pct_uber']:.1f}% de las ventas — "
            f"canal en crecimiento sostenido. Asegurar disponibilidad de top 20 productos en plataforma."
        )

    # Marca concentración
    if kpis["marca_1_pct"] > 20:
        insights.append(
            f"⚠️  {kpis['marca_1']} concentra el {kpis['marca_1_pct']:.1f}% de ventas. "
            f"Dependencia alta — evaluar desarrollo de marcas alternativas en categorías similares."
        )
    else:
        insights.append(
            f"✅ {kpis['marca_1']} lidera con {kpis['marca_1_pct']:.1f}% — "
            f"concentración saludable. Portafolio diversificado ({kpis['marcas_80']} marcas suman el 80%)."
        )

    # Día estrella
    insights.append(
        f"📅 Día {kpis['dia_max_num']} fue el más fuerte del mes "
        f"(${kpis['dia_max_venta']/1e3:.0f}K). "
        f"Analizar qué activación o evento lo provocó y planificar para replicarlo."
    )

    # Descuentos
    if kpis["pct_lineas_dto"] > 30:
        insights.append(
            f"🏷️ El {kpis['pct_lineas_dto']:.0f}% de líneas tiene descuento aplicado "
            f"(${kpis['monto_dto']/1e3:.0f}K cedidos). "
            f"Alta tasa — verificar si responde a política definida o es ad-hoc."
        )
    else:
        insights.append(
            f"✅ Política de descuentos controlada: {kpis['pct_lineas_dto']:.0f}% de líneas "
            f"con descuento (${kpis['monto_dto']/1e3:.0f}K). Margen protegido."
        )

    return insights[:6]


# ================================================================
#  PASO 5 — GENERAR PDF
# ================================================================
def generar_pdf(df_prod, kpis: dict, ctx: dict, insights: list,
                año: int, mes: int, output_path: str):

    mes_str = f"{MESES_ES[mes]} {año}"
    fig = plt.figure(figsize=(24, 30), facecolor=COLORES["fondo"])
    gs = gridspec.GridSpec(7, 4, figure=fig,
                           hspace=0.55, wspace=0.38,
                           top=0.935, bottom=0.03, left=0.05, right=0.97)

    # ── Header ─────────────────────────────────────────────────
    ax_h = fig.add_axes([0, 0.938, 1, 0.062])
    ax_h.set_facecolor(COLORES["primario"]); ax_h.axis("off")
    ax_h.text(0.03, 0.65, f"🌿 {NOMBRE_TIENDA}", fontsize=22,
              fontweight="bold", color="white", va="center")
    ax_h.text(0.03, 0.18, f"Reporte Mensual de Ventas — {mes_str}",
              fontsize=13, color=COLORES["claro"], va="center")
    ax_h.text(0.97, 0.50, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
              fontsize=9, color="#aaffcc", va="center", ha="right")

    # ── Helper: tarjeta KPI con delta ──────────────────────────
    def kpi_card(ax, titulo, valor, delta_txt=None, delta_pos=None, sub2=None):
        ax.set_facecolor(COLORES["neutro"]); ax.axis("off")
        rect = FancyBboxPatch((0.04, 0.04), 0.92, 0.92,
                              boxstyle="round,pad=0.02",
                              linewidth=1.2, edgecolor=COLORES["borde"],
                              facecolor=COLORES["neutro"])
        ax.add_patch(rect)
        ax.text(0.5, 0.84, titulo, fontsize=9, ha="center", color="#555555")
        ax.text(0.5, 0.52, valor, fontsize=16, ha="center",
                fontweight="bold", color=COLORES["primario"])
        if delta_txt:
            color_d = COLORES["positivo"] if delta_pos else COLORES["negativo"] if delta_pos is False else "#888"
            ax.text(0.5, 0.26, delta_txt, fontsize=9, ha="center", color=color_d)
        if sub2:
            ax.text(0.5, 0.10, sub2, fontsize=8, ha="center", color="#999999")

    # ── Helper: flecha delta ────────────────────────────────────
    def delta(actual, ref, pct=True, invertir=False):
        if ref is None or ref == 0: return None, None
        d = (actual - ref) / ref * 100 if pct else actual - ref
        pos = d >= 0 if not invertir else d <= 0
        sym = "▲" if d >= 0 else "▼"
        if pct:
            return f"{sym} {abs(d):.1f}% vs año ant.", pos
        else:
            return f"{sym} {abs(d):.1f}pp vs año ant.", pos

    # ── FILA 0: KPIs principales ────────────────────────────────
    d_vb, p_vb = delta(kpis["venta_bruta"], ctx.get("ly_venta"))
    kpi_card(fig.add_subplot(gs[0,0]), "Venta Bruta",
             f"${kpis['venta_bruta']/1e6:.2f}M", d_vb, p_vb,
             f"Neta: ${kpis['venta_neta']/1e6:.2f}M")

    d_mg, p_mg = delta(kpis["pct_margen"]*100, ctx.get("ly_margen"), pct=False)
    kpi_card(fig.add_subplot(gs[0,1]), "Margen Bruto",
             f"{kpis['pct_margen']*100:.1f}%", d_mg, p_mg,
             f"${kpis['margen_bruto']/1e6:.2f}M absoluto")

    d_bol, p_bol = delta(kpis["tickets"], ctx.get("ly_boletas"))
    kpi_card(fig.add_subplot(gs[0,2]), "Boletas",
             f"{kpis['tickets']:,}", d_bol, p_bol,
             f"Diario: {kpis['tickets']//kpis['dias_op']} boletas/día")

    d_bp, p_bp = delta(kpis["boleta_prom"], ctx.get("ly_bp"))
    kpi_card(fig.add_subplot(gs[0,3]), "Boleta Promedio",
             f"${kpis['boleta_prom']/1000:.0f}K", d_bp, p_bp,
             f"{kpis['u_ticket']:.2f} u/ticket")

    # ── FILA 1-2: Tendencia diaria + últimos 12 meses ──────────
    ax_tend = fig.add_subplot(gs[1:3, :3])
    pd_d = kpis["por_dia"]
    dias  = pd_d["dia"]
    vk    = pd_d["ventas"] / 1000
    cbar  = [COLORES["primario"] if v >= vk.mean() else COLORES["acento"] for v in vk]
    ax_tend.bar(dias, vk, color=cbar, width=0.75, zorder=3)
    ax_tend.plot(dias, vk.rolling(5, center=True).mean(),
                 color=COLORES["alerta"], linewidth=2.5, zorder=4, label="Media móvil 5d")
    ax_tend.axhline(vk.mean(), color="#aaa", linestyle="--", linewidth=1.2, label="Prom. mes")

    # Línea referencia mismo mes año anterior si existe
    if ctx.get("ly_venta") and kpis["dias_op"] > 0:
        ly_diaria = ctx["ly_venta"] / 1000 / kpis["dias_op"]
        ax_tend.axhline(ly_diaria, color=COLORES["azul"], linestyle=":",
                        linewidth=1.5, label=f"Prom. diario {año-1}")

    ax_tend.set_title(f"Tendencia Diaria — {mes_str} ($ miles CLP)",
                      fontsize=13, fontweight="bold", pad=12, color=COLORES["texto"])
    ax_tend.set_xlabel("Día del mes")
    ax_tend.set_ylabel("Ventas ($ miles)")
    ax_tend.legend(fontsize=8.5)
    ax_tend.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:.0f}K"))
    ax_tend.set_facecolor("#FAFCFA")
    ax_tend.grid(axis="y", alpha=0.3)
    ax_tend.spines[["top","right"]].set_visible(False)

    # ── FILA 1-2: Día de semana ────────────────────────────────
    ax_sem = fig.add_subplot(gs[1:3, 3])
    ds = kpis["por_dia_sem"]
    cv = [COLORES["primario"] if v == ds.max() else COLORES["acento"] for v in ds]
    ax_sem.barh(ds.index, ds.values/1000, color=cv)
    ax_sem.set_title("Ventas por\nDía de Semana", fontsize=11,
                     fontweight="bold", color=COLORES["texto"])
    ax_sem.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:.0f}K"))
    ax_sem.set_facecolor("#FAFCFA")
    ax_sem.spines[["top","right"]].set_visible(False)
    ax_sem.grid(axis="x", alpha=0.3)

    # ── FILA 3: Canales ────────────────────────────────────────
    ax_can = fig.add_subplot(gs[3, :2])
    canal = kpis["por_canal"].head(4)
    cmap_c = [COLORES["primario"], COLORES["acento"], COLORES["azul"], "#74C69D"]
    bars_c = ax_can.bar(range(len(canal)), canal.values/1000,
                        color=cmap_c[:len(canal)], width=0.6)
    ax_can.set_xticks(range(len(canal)))
    ax_can.set_xticklabels(canal.index, fontsize=9)
    ax_can.set_title("Ventas por Canal ($ miles)", fontsize=11,
                     fontweight="bold", color=COLORES["texto"])
    ax_can.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:.0f}K"))
    ax_can.set_facecolor("#FAFCFA")
    ax_can.spines[["top","right"]].set_visible(False)
    ax_can.grid(axis="y", alpha=0.3)
    for bar, val in zip(bars_c, canal.values):
        pct = val/kpis["venta_bruta"]*100
        ax_can.text(bar.get_x()+bar.get_width()/2, bar.get_height()+15,
                    f"${val/1e3:.0f}K\n({pct:.0f}%)", ha="center", va="bottom", fontsize=8)

    # ── FILA 3: Ultimos 12 meses ───────────────────────────────
    ax_12 = fig.add_subplot(gs[3, 2:])
    if ctx.get("ult12_ventas"):
        labs = [p[-5:] for p in ctx["ult12_periodos"]]
        vls  = [v/1e6 for v in ctx["ult12_ventas"]]
        ax_12.bar(range(len(vls)), vls,
                  color=[COLORES["secundario"]]*len(vls), alpha=0.7, width=0.7)
        ax_12.plot(range(len(vls)), vls, color=COLORES["primario"],
                   linewidth=2, marker="o", markersize=4, zorder=4)
        ax_12.set_xticks(range(len(labs)))
        ax_12.set_xticklabels(labs, rotation=45, ha="right", fontsize=7.5)
        ax_12.set_title("Últimos 12 Meses — Tendencia", fontsize=11,
                        fontweight="bold", color=COLORES["texto"])
        ax_12.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:.0f}M"))
        ax_12.set_facecolor("#FAFCFA")
        ax_12.spines[["top","right"]].set_visible(False)
        ax_12.grid(axis="y", alpha=0.3)
    else:
        ax_12.axis("off")
        ax_12.text(0.5, 0.5, "Sin historial disponible", ha="center",
                   va="center", color="#aaa", fontsize=11)

    # ── FILA 4-5: Top productos ────────────────────────────────
    ax_pr = fig.add_subplot(gs[4:6, :2])
    prods = kpis["top_prods"].sort_values()
    pc    = [COLORES["dorado"] if i == len(prods)-1 else COLORES["acento"]
             for i in range(len(prods))]
    labs_p = [p[:48]+"…" if len(p)>48 else p for p in prods.index]
    bars_p = ax_pr.barh(range(len(prods)), prods.values/1000, color=pc)
    ax_pr.set_yticks(range(len(prods)))
    ax_pr.set_yticklabels(labs_p, fontsize=7.5)
    ax_pr.set_title("Top 10 Productos del Mes ($ miles)", fontsize=12,
                    fontweight="bold", color=COLORES["texto"])
    ax_pr.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:.0f}K"))
    ax_pr.set_facecolor("#FAFCFA")
    ax_pr.spines[["top","right"]].set_visible(False)
    ax_pr.grid(axis="x", alpha=0.3)
    for bar, val in zip(bars_p, prods.values):
        ax_pr.text(bar.get_width()+2, bar.get_y()+bar.get_height()/2,
                   f"${val/1e3:.0f}K", va="center", fontsize=7.5)

    # ── FILA 4-5: Pareto marcas ────────────────────────────────
    ax_par = fig.add_subplot(gs[4:6, 2:])
    top10m = kpis["marcas_v"].head(10).sort_values()
    pm_c   = [COLORES["primario"] if i == len(top10m)-1 else COLORES["secundario"]
              for i in range(len(top10m))]
    ax_par.barh(range(len(top10m)), top10m.values/1000, color=pm_c)
    ax_par.set_yticks(range(len(top10m)))
    ax_par.set_yticklabels(top10m.index, fontsize=8)
    # Línea 80%
    acum = top10m[::-1].cumsum()
    total = kpis["marcas_v"].sum()
    for i, (marca, val) in enumerate(top10m[::-1].items()):
        pct_acum = acum.iloc[i] / total * 100
        if pct_acum >= 80:
            ax_par.axhline(len(top10m)-1-i-0.5, color=COLORES["alerta"],
                           linestyle="--", linewidth=1.5, label="Umbral 80%")
            break
    ax_par.set_title(f"Top 10 Marcas | {kpis['marcas_80']} marcas = 80% ventas",
                     fontsize=11, fontweight="bold", color=COLORES["texto"])
    ax_par.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"${x:.0f}K"))
    ax_par.set_facecolor("#FAFCFA")
    ax_par.spines[["top","right"]].set_visible(False)
    ax_par.grid(axis="x", alpha=0.3)
    ax_par.legend(fontsize=8.5)

    # ── FILA 6: Insights ───────────────────────────────────────
    ax_ins = fig.add_subplot(gs[6, :])
    ax_ins.set_facecolor(COLORES["claro"]); ax_ins.axis("off")
    rect3 = FancyBboxPatch((0.005, 0.04), 0.990, 0.92,
                           boxstyle="round,pad=0.01",
                           linewidth=1.5, edgecolor=COLORES["secundario"],
                           facecolor=COLORES["claro"])
    ax_ins.add_patch(rect3)
    ax_ins.text(0.012, 0.91, f"🔍 Insights Estratégicos — {mes_str}",
                fontsize=11, fontweight="bold", color=COLORES["primario"], va="top")
    col_w = 0.485
    for i, insight in enumerate(insights):
        col = i % 2; fila = i // 2
        x = 0.012 + col * col_w
        y = 0.74 - fila * 0.24
        ax_ins.text(x, y, insight, fontsize=8.3, color=COLORES["texto"],
                    va="top", wrap=True, transform=ax_ins.transAxes)

    plt.savefig(output_path, dpi=150, bbox_inches="tight",
                facecolor=COLORES["fondo"], format="pdf")
    plt.close()
    print(f"  📊 PDF guardado: {output_path}")


# ================================================================
#  PASO 6 — ACTUALIZAR MATRIZ MAESTRA
# ================================================================
def actualizar_matriz(año: int, mes: int):
    """Llama al script de consolidación para agregar el nuevo mes."""
    print(f"\n🔄 Actualizando matriz maestra...")
    import subprocess
    result = subprocess.run(
        ["python3", os.path.expanduser("~/Documents/Ciencia de Datos/rumbo-verde-analisis/scripts/consolidar_matriz.py")],
        capture_output=True, text=True
    )
    if result.returncode == 0:
        print("  ✅ Matriz actualizada correctamente")
    else:
        print(f"  ⚠️  Error al actualizar matriz: {result.stderr[:200]}")


# ================================================================
#  MAIN
# ================================================================
def main():
    if len(sys.argv) < 2:
        print("\n❗ Uso: python generar_reporte_mes.py <archivo_mes.xlsx>")
        print("   Ejemplo: python generar_reporte_mes.py mensuales/2026-06_ventas.xlsx\n")
        sys.exit(1)

    archivo = sys.argv[1]
    if not os.path.exists(archivo):
        print(f"\n❌ Archivo no encontrado: {archivo}")
        sys.exit(1)

    os.makedirs(CARPETA_REPORTES, exist_ok=True)
    os.makedirs(CARPETA_MENSUALES, exist_ok=True)

    print("=" * 60)
    print("  RUMBO VERDE — Generador de Reporte Mensual")
    print("=" * 60)

    # 1. Cargar mes
    df_prod, df_desp, df_devs, año, mes = cargar_mes(archivo)
    mes_str = f"{MESES_ES[mes]} {año}"

    # 2. Copiar a carpeta mensuales si no está ahí
    dest = os.path.join(CARPETA_MENSUALES, os.path.basename(archivo))
    if os.path.abspath(archivo) != os.path.abspath(dest):
        import shutil
        shutil.copy2(archivo, dest)
        print(f"  📁 Copiado a {dest}")

    # 3. Actualizar matriz
    actualizar_matriz(año, mes)

    # 4. Calcular KPIs del mes
    print(f"\n⚙️  Calculando KPIs de {mes_str}...")
    kpis = calcular_kpis_mes(df_prod, df_desp, df_devs)

    # 5. Leer contexto histórico
    print("📚 Leyendo contexto histórico...")
    ctx = leer_contexto_historico(año, mes)

    # 6. Generar insights
    insights = generar_insights(kpis, ctx, año, mes)

    # 7. Generar PDF
    base = f"reporte_{año}_{mes:02d}"
    pdf_path = os.path.join(CARPETA_REPORTES, f"{base}.pdf")
    print(f"\n🎨 Generando PDF...")
    generar_pdf(df_prod, kpis, ctx, insights, año, mes, pdf_path)

    # 8. Resumen en consola
    yoy_txt = ""
    if ctx.get("ly_venta"):
        yoy = (kpis["venta_bruta"] - ctx["ly_venta"]) / ctx["ly_venta"] * 100
        yoy_txt = f"YoY: {'+'if yoy>=0 else ''}{yoy:.1f}%"

    print(f"""
╔══════════════════════════════════════════════════════╗
║  ✅ REPORTE GENERADO — {mes_str:<28}║
╠══════════════════════════════════════════════════════╣
║  💰 Venta Bruta:    ${kpis['venta_bruta']/1e6:.2f}M  {yoy_txt:<20}║
║  📊 Margen:         {kpis['pct_margen']*100:.1f}%{'':<33}║
║  🛒 Boletas:        {kpis['tickets']:>6,}{'':<28}║
║  🧾 Boleta Prom.:   ${kpis['boleta_prom']/1000:.0f}K{'':<31}║
╠══════════════════════════════════════════════════════╣
║  📄 PDF:  {pdf_path:<42}║
╚══════════════════════════════════════════════════════╝
""")


if __name__ == "__main__":
    main()
