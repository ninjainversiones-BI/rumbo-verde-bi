"""
================================================================
  RUMBO VERDE — SCRIPT DE CONSOLIDACIÓN MATRIZ MAESTRA
  Construye una fila por mes con todos los KPIs definidos
================================================================

ESTRUCTURA DE CARPETAS EN TU MAC:
    ventas_rumbo_verde/
    ├── anuales/
    │   ├── 2022_ventas.xlsx
    │   ├── 2023_ventas.xlsx
    │   ├── 2024_ventas.xlsx
    │   └── 2025_ventas.xlsx
    ├── mensuales/
    │   ├── 2026-01_ventas.xlsx
    │   └── ...
    └── consolidar_matriz.py  ← este script

USO:
    python consolidar_matriz.py

OUTPUT:
    matriz_maestra.xlsx  → una fila por mes, 45 columnas de KPIs
================================================================
"""

import os
import re
import pandas as pd
import numpy as np
from datetime import datetime

# ================================================================
#  CONFIGURACIÓN — ajusta estas rutas si es necesario
# ================================================================
CARPETA_ANUALES   = os.path.expanduser("~/Documents/Rumbo Verde/Ventas/anuales")
CARPETA_MENSUALES = os.path.expanduser("~/Documents/Rumbo Verde/Ventas/mensuales")
OUTPUT_PATH       = os.path.expanduser("~/Documents/Ciencia de Datos/rumbo-verde-analisis/matriz/matriz_maestra.xlsx")

# ── Mapeo de canales ─────────────────────────────────────────────
# Normaliza todos los nombres históricos de listas de precio
# a 4 canales limpios. Agrega aquí si aparecen nuevos.
CANAL_MAP = {
    # Tienda física / web propia
    "lista de precios base rv final":   "Tienda / Web",
    "lista de precios base rv ":        "Tienda / Web",
    "cafeteria rumbo verde":            "Tienda / Web",
    "3r":                               "Tienda / Web",
    "precio mayorista":                 "Mayorista",

    # Cornershop (era 2022–2023)
    "cornershop":                       "Cornershop",   # catch-all para variantes

    # Uber Eats (era 2024–2026)
    "uber":                             "Uber Eats",    # catch-all para variantes

    # Rappi
    "rappi":                            "Rappi",

    # Mercado Libre
    "ml full":                          "Mercado Libre",
}

def normalizar_canal(lista_precio: str) -> str:
    """Convierte cualquier nombre de lista de precio al canal normalizado."""
    if pd.isna(lista_precio):
        return "Tienda / Web"
    lp = str(lista_precio).lower().strip()
    for key, canal in CANAL_MAP.items():
        if key in lp:
            return canal
    return "Tienda / Web"  # default


# ================================================================
#  PASO 1 — LIMPIAR Y PREPARAR UN DATAFRAME CRUDO
# ================================================================
def preparar_df(df: pd.DataFrame, año_esperado: int = None) -> pd.DataFrame:
    """
    Limpia el dataframe crudo:
    - Parsea fechas
    - Filtra al año correcto (evita solapamiento entre archivos)
    - Separa productos de despachos
    - Separa ventas de devoluciones
    - Normaliza canales
    - Excluye márgenes anómalos (errores de carga en sistema)
    """
    df = df.copy()

    # Parsear fecha
    df["Fecha Venta"] = pd.to_datetime(
        df["Fecha Venta"], dayfirst=True, errors="coerce"
    )
    df = df.dropna(subset=["Fecha Venta"])
    df["año"]  = df["Fecha Venta"].dt.year
    df["mes"]  = df["Fecha Venta"].dt.month
    df["dia"]  = df["Fecha Venta"].dt.day

    # Filtrar al año esperado (evita filas de dic/ene del archivo vecino)
    if año_esperado:
        df = df[df["año"] == año_esperado]

    # Normalizar canal
    df["canal"] = df["Lista de Precio"].apply(normalizar_canal)

    # Separar tipos
    df_ventas    = df[df["Tipo Movimiento"] == "venta"].copy()
    df_devs      = df[df["Tipo Movimiento"] == "devolucion"].copy()

    # Separar despachos (costo = 0) de productos reales
    df_productos = df_ventas[df_ventas["Costo neto unitario"] > 0].copy()
    df_despachos = df_ventas[df_ventas["Costo neto unitario"] == 0].copy()

    # Excluir márgenes anómalos (errores de carga: costo > precio * 2)
    # Estos distorsionan el % margen del mes completo
    margen_anomalo = (
        (df_productos["Costo neto unitario"] > df_productos["Precio Bruto Unitario"] * 2) &
        (df_productos["Precio Bruto Unitario"] < 100)
    )
    df_productos = df_productos[~margen_anomalo].copy()

    return df_productos, df_despachos, df_devs


# ================================================================
#  PASO 2 — CALCULAR KPIs DE UN MES
# ================================================================
def calcular_mes(df_prod, df_desp, df_devs, año, mes) -> dict:
    """
    Recibe los dataframes ya filtrados para un mes específico
    y devuelve un diccionario con todos los KPIs de la matriz.
    """

    # ── Filtrar al mes ──────────────────────────────────────────
    prod  = df_prod[(df_prod["año"] == año) & (df_prod["mes"] == mes)]
    desp  = df_desp[(df_desp["año"] == año) & (df_desp["mes"] == mes)]
    devs  = df_devs[(df_devs["año"] == año) & (df_devs["mes"] == mes)]

    if len(prod) == 0:
        return None

    meses_es = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",
                6:"Junio",7:"Julio",8:"Agosto",9:"Septiembre",
                10:"Octubre",11:"Noviembre",12:"Diciembre"}

    # ── Métricas base ───────────────────────────────────────────
    venta_bruta  = prod["Venta Total Bruta"].sum()
    venta_neta   = prod["Venta Total Neta"].sum()
    margen_bruto = prod["Margen"].sum()
    pct_margen   = margen_bruto / venta_neta if venta_neta > 0 else np.nan

    tickets      = prod["Tracking number"].nunique()
    unidades     = prod["Cantidad"].sum()

    boleta_prom        = venta_bruta / tickets if tickets > 0 else np.nan
    precio_art_prom    = venta_bruta / unidades if unidades > 0 else np.nan
    unidades_por_ticket = unidades / tickets if tickets > 0 else np.nan

    dias_operativos    = prod["dia"].nunique()
    venta_diaria_prom  = venta_bruta / dias_operativos if dias_operativos > 0 else np.nan

    # Día de mayor venta
    venta_x_dia  = prod.groupby("dia")["Venta Total Bruta"].sum()
    dia_mayor    = int(venta_x_dia.idxmax()) if len(venta_x_dia) > 0 else np.nan
    venta_mayor  = venta_x_dia.max() if len(venta_x_dia) > 0 else np.nan

    # ── Descuentos ──────────────────────────────────────────────
    monto_descuentos       = prod["Descuento Bruto"].sum()
    lineas_con_dto         = (prod["% Descuento"] > 0).sum()
    pct_lineas_con_dto     = lineas_con_dto / len(prod) * 100 if len(prod) > 0 else np.nan
    descuento_prom_boleta  = monto_descuentos / tickets if tickets > 0 else np.nan

    # ── Despachos ───────────────────────────────────────────────
    ingresos_despacho     = desp["Venta Total Bruta"].sum()
    boletas_con_despacho  = desp["Tracking number"].nunique()

    # ── Devoluciones ────────────────────────────────────────────
    monto_devoluciones = abs(devs["Venta Total Bruta"].sum()) if len(devs) > 0 else 0

    # ── Canales ─────────────────────────────────────────────────
    por_canal = prod.groupby("canal")["Venta Total Bruta"].sum()

    venta_tienda   = por_canal.get("Tienda / Web", 0)
    venta_uber     = por_canal.get("Uber Eats", 0)
    venta_cs       = por_canal.get("Cornershop", 0)
    venta_ml       = por_canal.get("Mercado Libre", 0)
    venta_rappi    = por_canal.get("Rappi", 0)
    venta_mayorista= por_canal.get("Mayorista", 0)

    pct_tienda = venta_tienda / venta_bruta * 100 if venta_bruta > 0 else np.nan
    pct_uber   = venta_uber   / venta_bruta * 100 if venta_bruta > 0 else np.nan
    pct_cs     = venta_cs     / venta_bruta * 100 if venta_bruta > 0 else np.nan

    # ── Clientes ────────────────────────────────────────────────
    # Solo disponible cuando hay datos de cliente registrados
    clientes_df = prod[
        prod["Nombre Cliente"].notna() &
        (~prod["Nombre Cliente"].str.strip().str.lower().isin(["sin cliente", ""]))
    ]
    tiene_clientes = len(clientes_df) > 0

    if tiene_clientes:
        clientes_unicos      = clientes_df["Nombre Cliente"].nunique()
        # Frecuencia = tickets con cliente / clientes únicos
        tickets_con_cliente  = clientes_df["Tracking number"].nunique()
        frecuencia_compra    = tickets_con_cliente / clientes_unicos if clientes_unicos > 0 else np.nan
    else:
        clientes_unicos   = np.nan
        frecuencia_compra = np.nan

    # ── Pareto Marcas ────────────────────────────────────────────
    marcas_venta = (prod.groupby("Marca")["Venta Total Bruta"]
                    .sum().sort_values(ascending=False))
    marcas_venta = marcas_venta[marcas_venta.index.notna()]

    total_marcas_activas = len(marcas_venta)
    marca_1_nombre = marcas_venta.index[0] if len(marcas_venta) > 0 else np.nan
    marca_1_venta  = marcas_venta.iloc[0]  if len(marcas_venta) > 0 else np.nan
    marca_1_pct    = marca_1_venta / venta_bruta * 100 if venta_bruta > 0 else np.nan

    # Cuántas marcas suman el 80%
    acumulado = marcas_venta.cumsum() / marcas_venta.sum()
    marcas_80 = int((acumulado <= 0.80).sum()) + 1

    # % que concentran esas marcas_80 marcas
    pct_concentracion_pareto = marcas_venta.iloc[:marcas_80].sum() / venta_bruta * 100 if venta_bruta > 0 else np.nan

    # Top 3 marcas para referencia rápida
    top3_marcas_pct = marcas_venta.iloc[:3].sum() / venta_bruta * 100 if venta_bruta > 0 else np.nan

    # ── Pareto Productos ─────────────────────────────────────────
    prods_venta = (prod.groupby("Producto / Servicio")["Venta Total Bruta"]
                   .sum().sort_values(ascending=False))

    total_productos_activos = len(prods_venta)
    producto_estrella       = prods_venta.index[0] if len(prods_venta) > 0 else np.nan
    producto_estrella_venta = prods_venta.iloc[0]  if len(prods_venta) > 0 else np.nan
    producto_estrella_pct   = producto_estrella_venta / venta_bruta * 100 if venta_bruta > 0 else np.nan

    acum_prod = prods_venta.cumsum() / prods_venta.sum()
    productos_80 = int((acum_prod <= 0.80).sum()) + 1

    # ── Construir fila ───────────────────────────────────────────
    return {
        # Tiempo
        "periodo":              f"{año}-{mes:02d}",
        "año":                  año,
        "mes":                  mes,
        "mes_nombre":           meses_es[mes],
        "dias_operativos":      dias_operativos,

        # Ventas
        "venta_bruta":          round(venta_bruta, 0),
        "venta_neta":           round(venta_neta, 0),
        "venta_diaria_promedio":round(venta_diaria_prom, 0),
        "crecimiento_mom":      np.nan,   # se calcula al final
        "crecimiento_yoy":      np.nan,   # se calcula al final

        # Rentabilidad
        "margen_bruto":         round(margen_bruto, 0),
        "pct_margen":           round(pct_margen * 100, 2) if not np.isnan(pct_margen) else np.nan,
        "crecimiento_margen_yoy": np.nan, # se calcula al final

        # Descuentos
        "monto_descuentos":         round(monto_descuentos, 0),
        "pct_lineas_con_descuento": round(pct_lineas_con_dto, 1),
        "descuento_promedio_boleta":round(descuento_prom_boleta, 0),

        # Operación
        "boletas":                  tickets,
        "boleta_promedio":          round(boleta_prom, 0),
        "unidades_totales":         int(unidades),
        "precio_articulo_promedio": round(precio_art_prom, 0),
        "unidades_por_ticket":      round(unidades_por_ticket, 2),
        "dia_mayor_venta":          dia_mayor,
        "venta_dia_mayor":          round(venta_mayor, 0),

        # Despachos
        "ingresos_despacho":        round(ingresos_despacho, 0),
        "boletas_con_despacho":     boletas_con_despacho,

        # Devoluciones
        "monto_devoluciones":       round(monto_devoluciones, 0),

        # Canales
        "venta_bruta_tienda":   round(venta_tienda, 0),
        "venta_bruta_uber":     round(venta_uber, 0),
        "venta_bruta_cornershop":round(venta_cs, 0),
        "venta_bruta_ml":       round(venta_ml, 0),
        "venta_bruta_rappi":    round(venta_rappi, 0),
        "venta_bruta_mayorista":round(venta_mayorista, 0),
        "pct_tienda":           round(pct_tienda, 1) if not np.isnan(pct_tienda) else np.nan,
        "pct_uber":             round(pct_uber, 1) if not np.isnan(pct_uber) else np.nan,
        "pct_cornershop":       round(pct_cs, 1) if not np.isnan(pct_cs) else np.nan,

        # Clientes
        "clientes_unicos":          clientes_unicos if not np.isnan(clientes_unicos) else np.nan,
        "frecuencia_compra_promedio":round(frecuencia_compra, 2) if not np.isnan(frecuencia_compra) else np.nan,

        # Pareto Marcas
        "marca_1_nombre":               marca_1_nombre,
        "marca_1_pct":                  round(marca_1_pct, 1) if not np.isnan(marca_1_pct) else np.nan,
        "marcas_para_80pct":            marcas_80,
        "total_marcas_activas":         total_marcas_activas,
        "pct_concentracion_top3_marcas":round(top3_marcas_pct, 1),
        "pct_concentracion_pareto":     round(pct_concentracion_pareto, 1),

        # Pareto Productos
        "producto_estrella":            producto_estrella,
        "producto_estrella_pct":        round(producto_estrella_pct, 1) if not np.isnan(producto_estrella_pct) else np.nan,
        "productos_para_80pct":         productos_80,
        "total_productos_activos":      total_productos_activos,
    }


# ================================================================
#  PASO 3 — CALCULAR CRECIMIENTOS (MoM y YoY)
# ================================================================
def calcular_crecimientos(df: pd.DataFrame) -> pd.DataFrame:
    """
    Una vez que la matriz tiene todos los meses,
    calcula los crecimientos relativos mes a mes y año a año.
    """
    df = df.sort_values(["año", "mes"]).reset_index(drop=True)

    # MoM: comparación con el mes anterior
    df["crecimiento_mom"] = df["venta_bruta"].pct_change() * 100

    # YoY: mismo mes del año anterior
    df_yoy = df[["año", "mes", "venta_bruta", "margen_bruto"]].copy()
    df_yoy["año"] = df_yoy["año"] + 1
    df_yoy = df_yoy.rename(columns={
        "venta_bruta":  "venta_bruta_ly",
        "margen_bruto": "margen_bruto_ly"
    })
    df = df.merge(df_yoy, on=["año", "mes"], how="left")

    df["crecimiento_yoy"] = (
        (df["venta_bruta"] - df["venta_bruta_ly"]) / df["venta_bruta_ly"] * 100
    )
    df["crecimiento_margen_yoy"] = (
        (df["margen_bruto"] - df["margen_bruto_ly"]) / df["margen_bruto_ly"].abs() * 100
    )

    df = df.drop(columns=["venta_bruta_ly", "margen_bruto_ly"])

    # Redondear crecimientos
    for col in ["crecimiento_mom", "crecimiento_yoy", "crecimiento_margen_yoy"]:
        df[col] = df[col].round(1)

    return df


# ================================================================
#  PASO 4 — LEER ARCHIVOS Y CONSTRUIR LA MATRIZ
# ================================================================
def leer_archivo_anual(path: str, año: int) -> tuple:
    """Lee un archivo anual y retorna los 3 dataframes limpios."""
    print(f"  📂 Leyendo {os.path.basename(path)}...")
    df_raw = pd.read_excel(path)
    return preparar_df(df_raw, año_esperado=año)


def leer_archivo_mensual(path: str) -> tuple:
    """
    Lee un archivo mensual. Detecta el año desde los datos
    (no desde el nombre del archivo, para ser robusto).
    """
    print(f"  📂 Leyendo {os.path.basename(path)}...")
    df_raw = pd.read_excel(path)
    df_raw["Fecha Venta"] = pd.to_datetime(
        df_raw["Fecha Venta"], dayfirst=True, errors="coerce"
    )
    # Detectar año y mes predominante
    año_predominante = df_raw["Fecha Venta"].dt.year.mode()[0]
    return preparar_df(df_raw, año_esperado=int(año_predominante))


def construir_matriz() -> pd.DataFrame:
    """
    Función principal. Lee todos los archivos disponibles
    y construye la matriz maestra completa.
    """
    todas_las_filas = []

    # ── Archivos anuales ────────────────────────────────────────
    if os.path.exists(CARPETA_ANUALES):
        archivos_anuales = sorted([
            f for f in os.listdir(CARPETA_ANUALES)
            if f.endswith(".xlsx") or f.endswith(".xls")
        ])
        if archivos_anuales:
            print(f"\n📁 Procesando {len(archivos_anuales)} archivo(s) anual(es)...")
        for archivo in archivos_anuales:
            path = os.path.join(CARPETA_ANUALES, archivo)
            # Extraer año del nombre del archivo (ej: 2024_ventas.xlsx)
            match = re.search(r"(20\d{2})", archivo)
            if not match:
                print(f"  ⚠️  No se pudo detectar año en: {archivo} — saltando")
                continue
            año = int(match.group(1))
            prod, desp, devs = leer_archivo_anual(path, año)

            meses_presentes = sorted(prod["mes"].unique())
            for mes in meses_presentes:
                fila = calcular_mes(prod, desp, devs, año, mes)
                if fila:
                    todas_las_filas.append(fila)
                    print(f"    ✅ {año}-{mes:02d} → {fila['boletas']:,} boletas | ${fila['venta_bruta']/1e6:.2f}M")

    # ── Archivos mensuales ──────────────────────────────────────
    if os.path.exists(CARPETA_MENSUALES):
        archivos_mensuales = sorted([
            f for f in os.listdir(CARPETA_MENSUALES)
            if f.endswith(".xlsx") or f.endswith(".xls")
        ])
        if archivos_mensuales:
            print(f"\n📁 Procesando {len(archivos_mensuales)} archivo(s) mensual(es)...")
        for archivo in archivos_mensuales:
            path = os.path.join(CARPETA_MENSUALES, archivo)
            prod, desp, devs = leer_archivo_mensual(path)
            if len(prod) == 0:
                print(f"  ⚠️  Sin datos en: {archivo}")
                continue
            año = int(prod["año"].mode()[0])
            mes = int(prod["mes"].mode()[0])
            fila = calcular_mes(prod, desp, devs, año, mes)
            if fila:
                todas_las_filas.append(fila)
                print(f"    ✅ {año}-{mes:02d} → {fila['boletas']:,} boletas | ${fila['venta_bruta']/1e6:.2f}M")

    if not todas_las_filas:
        print("\n❌ No se encontraron datos. Verifica las carpetas 'anuales' y 'mensuales'.")
        return pd.DataFrame()

    # ── Consolidar ──────────────────────────────────────────────
    matriz = pd.DataFrame(todas_las_filas)
    matriz = matriz.sort_values(["año", "mes"]).reset_index(drop=True)

    # Evitar duplicados (si un mes aparece en anual y mensual)
    matriz = matriz.drop_duplicates(subset=["año", "mes"], keep="last")

    # Calcular crecimientos
    matriz = calcular_crecimientos(matriz)

    return matriz


# ================================================================
#  PASO 5 — EXPORTAR A EXCEL CON FORMATO
# ================================================================
def exportar_excel(matriz: pd.DataFrame, output_path: str):
    """Guarda la matriz en Excel con formato profesional."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

    wb = Workbook()
    ws = wb.active
    ws.title = "Matriz Maestra"
    ws.sheet_view.showGridLines = False

    VERDE    = "1B4332"
    VERDE2   = "40916C"
    CLARO    = "D8F3DC"
    BLANCO   = "FFFFFF"
    GRIS     = "F8FAF8"

    # ── Grupos de columnas con colores ──────────────────────────
    grupos = {
        "Tiempo":       (["periodo","año","mes","mes_nombre","dias_operativos"], "2D6A4F"),
        "Ventas":       (["venta_bruta","venta_neta","venta_diaria_promedio",
                          "crecimiento_mom","crecimiento_yoy"], "1A759F"),
        "Rentabilidad": (["margen_bruto","pct_margen","crecimiento_margen_yoy"], "168AAD"),
        "Descuentos":   (["monto_descuentos","pct_lineas_con_descuento",
                          "descuento_promedio_boleta"], "E76F51"),
        "Operación":    (["boletas","boleta_promedio","unidades_totales",
                          "precio_articulo_promedio","unidades_por_ticket",
                          "dia_mayor_venta","venta_dia_mayor"], "52B788"),
        "Despachos":    (["ingresos_despacho","boletas_con_despacho"], "74C69D"),
        "Devoluciones": (["monto_devoluciones"], "E9C46A"),
        "Canales":      (["venta_bruta_tienda","venta_bruta_uber","venta_bruta_cornershop",
                          "venta_bruta_ml","venta_bruta_rappi","venta_bruta_mayorista",
                          "pct_tienda","pct_uber","pct_cornershop"], "264653"),
        "Clientes":     (["clientes_unicos","frecuencia_compra_promedio"], "2A9D8F"),
        "Pareto Marcas":(["marca_1_nombre","marca_1_pct","marcas_para_80pct",
                          "total_marcas_activas","pct_concentracion_top3_marcas",
                          "pct_concentracion_pareto"], "F4A261"),
        "Pareto Prod.": (["producto_estrella","producto_estrella_pct",
                          "productos_para_80pct","total_productos_activos"], "E76F51"),
    }

    # Orden de columnas según grupos
    orden_cols = []
    for grupo, (cols, _) in grupos.items():
        for c in cols:
            if c in matriz.columns:
                orden_cols.append(c)

    # Asegurar que no falte ninguna columna
    for c in matriz.columns:
        if c not in orden_cols:
            orden_cols.append(c)

    matriz = matriz[orden_cols]

    # ── Fila 1: título ──────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(orden_cols))}1")
    ws["A1"] = f"🌿 RUMBO VERDE — Matriz Maestra de KPIs | Generada: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(bold=True, size=13, color=BLANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=VERDE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Fila 2: nombres de grupo ────────────────────────────────
    col_idx = 1
    for grupo, (cols, color) in grupos.items():
        cols_presentes = [c for c in cols if c in matriz.columns]
        if not cols_presentes:
            continue
        inicio = col_idx
        fin    = col_idx + len(cols_presentes) - 1
        if inicio == fin:
            cell = ws.cell(2, inicio, grupo)
        else:
            ws.merge_cells(
                start_row=2, start_column=inicio,
                end_row=2,   end_column=fin
            )
            cell = ws.cell(2, inicio)
            cell.value = grupo
        cell.font      = Font(bold=True, size=9, color=BLANCO)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += len(cols_presentes)
    ws.row_dimensions[2].height = 18

    # ── Fila 3: headers de columnas ─────────────────────────────
    headers_es = {
        "periodo": "Período", "año": "Año", "mes": "Mes", "mes_nombre": "Mes Nombre",
        "dias_operativos": "Días Operativos",
        "venta_bruta": "Venta Bruta ($)", "venta_neta": "Venta Neta ($)",
        "venta_diaria_promedio": "Venta Diaria Prom ($)",
        "crecimiento_mom": "Crec. MoM (%)", "crecimiento_yoy": "Crec. YoY (%)",
        "margen_bruto": "Margen Bruto ($)", "pct_margen": "% Margen",
        "crecimiento_margen_yoy": "Crec. Margen YoY (%)",
        "monto_descuentos": "Dctos. Entregados ($)",
        "pct_lineas_con_descuento": "% Líneas con Dcto.",
        "descuento_promedio_boleta": "Dcto. Prom/Boleta ($)",
        "boletas": "Boletas", "boleta_promedio": "Boleta Promedio ($)",
        "unidades_totales": "Unidades Totales",
        "precio_articulo_promedio": "Precio Art. Prom ($)",
        "unidades_por_ticket": "Unidades/Ticket",
        "dia_mayor_venta": "Día Mayor Venta", "venta_dia_mayor": "Venta Día Mayor ($)",
        "ingresos_despacho": "Ingresos Despacho ($)", "boletas_con_despacho": "Boletas c/Despacho",
        "monto_devoluciones": "Devoluciones ($)",
        "venta_bruta_tienda": "Venta Tienda ($)", "venta_bruta_uber": "Venta Uber ($)",
        "venta_bruta_cornershop": "Venta Cornershop ($)", "venta_bruta_ml": "Venta ML ($)",
        "venta_bruta_rappi": "Venta Rappi ($)", "venta_bruta_mayorista": "Venta Mayorista ($)",
        "pct_tienda": "% Tienda", "pct_uber": "% Uber", "pct_cornershop": "% Cornershop",
        "clientes_unicos": "Clientes Únicos", "frecuencia_compra_promedio": "Frec. Compra Prom.",
        "marca_1_nombre": "Marca #1", "marca_1_pct": "% Marca #1",
        "marcas_para_80pct": "Marcas p/80%", "total_marcas_activas": "Total Marcas",
        "pct_concentracion_top3_marcas": "% Conc. Top3 Marcas",
        "pct_concentracion_pareto": "% Conc. Pareto",
        "producto_estrella": "Producto Estrella", "producto_estrella_pct": "% Prod. Estrella",
        "productos_para_80pct": "Productos p/80%", "total_productos_activos": "Total Productos",
    }

    for j, col in enumerate(orden_cols, 1):
        cell = ws.cell(3, j, headers_es.get(col, col))
        cell.font      = Font(bold=True, size=9, color=BLANCO)
        cell.fill      = PatternFill("solid", fgColor=VERDE2)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 32

    # ── Datos ───────────────────────────────────────────────────
    for i, (_, row) in enumerate(matriz.iterrows(), 4):
        bg = PatternFill("solid", fgColor=GRIS) if i % 2 == 0 else None
        for j, col in enumerate(orden_cols, 1):
            val = row[col]
            cell = ws.cell(i, j)
            # Convertir numpy types
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not np.isnan(val) else None
            cell.value = val
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if bg:
                cell.fill = bg
            cell.font = Font(size=9)

    # ── Anchos de columna ───────────────────────────────────────
    anchos = {
        "periodo": 10, "año": 6, "mes": 5, "mes_nombre": 11, "dias_operativos": 8,
        "marca_1_nombre": 22, "producto_estrella": 38,
    }
    for j, col in enumerate(orden_cols, 1):
        ancho = anchos.get(col, 16)
        ws.column_dimensions[get_column_letter(j)].width = ancho

    # ── Congelar paneles ────────────────────────────────────────
    ws.freeze_panes = "F4"  # Congela columnas de tiempo

    wb.save(output_path)
    print(f"\n✅ Matriz guardada en: {output_path}")


# ================================================================
#  MAIN
# ================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("  RUMBO VERDE — Consolidación Matriz Maestra")
    print("=" * 60)

    matriz = construir_matriz()

    if len(matriz) == 0:
        exit(1)

    exportar_excel(matriz, OUTPUT_PATH)

    # Resumen en consola
    print(f"""
┌─────────────────────────────────────────────────────┐
│  📊 MATRIZ GENERADA EXITOSAMENTE                    │
├─────────────────────────────────────────────────────┤
│  Meses procesados : {len(matriz):<33}│
│  Período          : {str(matriz['periodo'].iloc[0]):<10} → {str(matriz['periodo'].iloc[-1]):<18}│
│  Columnas KPI     : {len(matriz.columns):<33}│
│  Archivo output   : {OUTPUT_PATH:<33}│
└─────────────────────────────────────────────────────┘
""")

    # Vista previa
    print("Vista previa (primeros 3 meses):")
    preview_cols = ["periodo", "venta_bruta", "pct_margen", "boletas",
                    "boleta_promedio", "crecimiento_yoy", "marca_1_nombre"]
    preview_cols = [c for c in preview_cols if c in matriz.columns]
    print(matriz[preview_cols].head(3).to_string(index=False))
